from statistics import median
import cv2
import numpy as np
import matplotlib.pyplot as plt
from itertools import combinations
import itertools
import os
import random
from scipy.spatial import Delaunay, ConvexHull
from math import degrees
from skimage.metrics import structural_similarity as ssim
from skimage.metrics import peak_signal_noise_ratio as psnr
from matplotlib.backend_bases import MouseEvent
from sympy.codegen.ast import Return
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl import Workbook, load_workbook


# # 创建Excel文件记录结果
excel_path = 'Try_.xlsx'

try:
    wb = load_workbook(excel_path)  # 如果文件已存在则加载
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active

#不同参数设置
config_list = [

    (7, 20, 22, 'HDU_W_20.png'),
    (6, 24, 19, 'HDU_W_24.png'),
    (5, 32, 15, 'HDU_W_32.png'),
    (4, 38, 13, 'HDU_W_38.png'),

    # (7, 16, 26, 'HDU_W_16.png'),


]

# 主次梯度阈值
energy_threshold =12
mean_size_2=1
Channel=0

# 要批量处理的载体图像文件
# image_list = ['.bmp',]
image_list = ['cactusfl.bmp','ivytree.bmp','utahmtn.bmp','clinmill.bmp','peppers.bmp','bodie.bmp','bluheron.bmp','baboon.bmp',]  # 你可以替换为目标图像列表


modes = [

    # 鲁棒性实验的多种攻击
    # ("ini",lambda img: img),
    # ("jpeg90", lambda img,output_path: cv2.imread(f"{output_folder}/{mode_name}_output.jpg") if cv2.imwrite(f"{output_folder}/{mode_name}_output.jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 90]) else img),
    ("jpeg70", lambda img,output_path: cv2.imread(f"{output_folder}/{mode_name}_J7_output.jpg") if cv2.imwrite( f"{output_folder}/{mode_name}_J7_output.jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 70]) else img),
    ("jpeg50", lambda img,output_path: cv2.imread(f"{output_folder}/{mode_name}_J5_output.jpg") if cv2.imwrite(
        f"{output_folder}/{mode_name}_J5_output.jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 50]) else img),
    ("jpeg30", lambda img, output_path: cv2.imread(f"{output_folder}/{mode_name}_J3_output.jpg") if cv2.imwrite(
        f"{output_folder}/{mode_name}_J3_output.jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 30]) else img),

    ("scale0.7", lambda img,output_path: cv2.resize(cv2.resize(img, (int(img.shape[1] * 0.7), int(img.shape[0] * 0.7)), interpolation=cv2.INTER_LINEAR), (img.shape[1], img.shape[0]), interpolation=cv2.INTER_LINEAR)),

         ("scale0.5",
     lambda img,output_path: cv2.resize(cv2.resize(img, (img.shape[1] // 2, img.shape[0] // 2), interpolation=cv2.INTER_LINEAR),
                            (img.shape[1], img.shape[0]), interpolation=cv2.INTER_LINEAR)),

     # ("scale1.3", lambda img,output_path: cv2.resize(
     #      cv2.resize(img, (int(img.shape[1] * 1.3), int(img.shape[0] * 1.3)), interpolation=cv2.INTER_LINEAR),
     #      (img.shape[1], img.shape[0]), interpolation=cv2.INTER_LINEAR)),
    ("scale1.5", lambda img,output_path: cv2.resize(
          cv2.resize(img, (int(img.shape[1] * 1.5), int(img.shape[0] * 1.5)), interpolation=cv2.INTER_LINEAR),
          (img.shape[1], img.shape[0]), interpolation=cv2.INTER_LINEAR)),
    # ("rotate90", lambda img, output_path: rotate_image(img, angle=90)),
    ("salt_pepper0.001", lambda img,output_path: add_salt_and_pepper_noise(img,  amount=0.001)),
     ("salt_pepper0.01", lambda img,output_path: add_salt_and_pepper_noise(img,  amount=0.01)),
     # ("gaussian_noise0.001", lambda img,output_path: add_gaussian_noise(img,sigma=0.001)),
     ("gaussian_noise0.01", lambda img,output_path: add_gaussian_noise(img,sigma=0.01)),
    ("gaussian_noise0.03", lambda img,output_path: add_gaussian_noise(img,sigma=0.03)),
    ("median_blur3", lambda img,output_path: apply_median_blur(img,strength=3)),
    ("mean_blur3", lambda img,output_path: apply_mean_filter(img,kernel_size=3)),

    ("scale0.8_jpeg80", lambda img,output_path: apply_scale_then_jpeg(img, scale_ratio=0.8, jpeg_quality=80,
                                                      output_path=f"{output_folder}/{mode_name}_S8J8_.jpg")),
    ("scale0.7_jpeg70", lambda img,output_path: apply_scale_then_jpeg(img, scale_ratio=0.7, jpeg_quality=70,
                                                      output_path=f"{output_folder}/{mode_name}_S7J7_.jpg")),

    ("noise1_then_jpeg8", lambda img,output_path: apply_noise_then_jpeg(img, sigma=0.01, jpeg_quality=80, output_path=f"{output_folder}/{mode_name}_noise1_jpeg8.jpg")),
    ("noise3_then_jpeg7", lambda img,output_path: apply_noise_then_jpeg(img, sigma=0.03, jpeg_quality=70,
                                                          output_path=f"{output_folder}/{mode_name}_noise3_jpeg7.jpg")),

    ("MB3_then_jpeg80", lambda img,output_path: apply_MB_then_jpeg(img, sigma=3, jpeg_quality=80, output_path=f"{output_folder}/{mode_name}_MB3_jpeg8.jpg")),
    ("MB3_then_jpeg50", lambda img,output_path: apply_MB_then_jpeg(img, sigma=3, jpeg_quality=50,
                                                          output_path=f"{output_folder}/{mode_name}_MB3_jpeg5.jpg")),

    ("scale8_then_noise1", lambda img,output_path: apply_scale_then_noise(img, scale_ratio=0.8, sigma=0.01)),
    ("scale7_then_noise3", lambda img,output_path: apply_scale_then_noise(img, scale_ratio=0.7, sigma=0.03)),
    ("scale_8noise01_then_jpeg8", lambda img,output_path: apply_scale_then_noise_then_jpeg(img, scale_ratio=0.8, sigma=0.01, jpeg_quality=80,
                                                          output_path=f"{output_folder}/{mode_name}_scale8_noise1_jpeg8.jpg")),
    ("scale_7noise_3then_jpeg7",
     lambda img,output_path: apply_scale_then_noise_then_jpeg(img, scale_ratio=0.7, sigma=0.03, jpeg_quality=70,
                                                  output_path=f"{output_folder}/{mode_name}_scale7_noise3_jpeg7.jpg")),
    ("scale_5noise_5then_jpeg5",
     lambda img,output_path: apply_scale_then_noise_then_jpeg(img, scale_ratio=0.5, sigma=0.05, jpeg_quality=50,
                                                  output_path=f"{output_folder}/{mode_name}_scale5_noise5_jpeg5.jpg")),
     ("scale_5noise_3then_jpeg5",
     lambda img,output_path: apply_scale_then_noise_then_jpeg(img, scale_ratio=0.5, sigma=0.03, jpeg_quality=50,
                                                  output_path=f"{output_folder}/{mode_name}_scale5_noise3_jpeg5.jpg")),

]


def sift_corner_detection(image, max_corners):
    sift = cv2.SIFT_create()
    keypoints = sift.detect(image, None)
    keypoints = sorted(keypoints, key=lambda k: -k.response)
    corners = np.array([kp.pt for kp in keypoints[:max_corners]], dtype=np.float32)
    return corners


def apply_median_blur(image, strength):

    # 确保 kernel size 是奇数且 ≥ 3
    ksize = max(3, strength)
    if ksize % 2 == 0:
        ksize += 1

    blurred = cv2.medianBlur(image, ksize)
    return blurred

def apply_mean_filter(image, kernel_size):

    if kernel_size % 2 == 0:
        kernel_size += 1  # 确保是奇数
    return cv2.blur(image, (kernel_size, kernel_size))

def apply_gaussian_blur(image, sigma):
    # 自动估算合适的核尺寸，必须为奇数
    ksize = int(6 * sigma + 1)
    if ksize % 2 == 0:
        ksize += 1

    blurred = cv2.GaussianBlur(image, (ksize, ksize), sigmaX=sigma, sigmaY=sigma)
    return blurred

def apply_mean_jpeg(image, kernel_size,jpeg_quality):
    if kernel_size % 2 == 0:
        kernel_size += 1  # 确保是奇数
    mean= cv2.blur(image, (kernel_size, kernel_size))
    success = cv2.imwrite(output_path, mean, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
    return cv2.imread(output_path) if success else mean

def rotate_image(image, angle):
    (h, w) = image.shape[:2]
    center = (w / 2, h / 2)

    # 计算旋转矩阵
    rot_mat = cv2.getRotationMatrix2D(center, angle, 1.0)

    # 计算新图像尺寸（使图像不裁剪）
    abs_cos = abs(rot_mat[0, 0])
    abs_sin = abs(rot_mat[0, 1])
    new_w = int(h * abs_sin + w * abs_cos)
    new_h = int(h * abs_cos + w * abs_sin)

    # 调整旋转矩阵中的平移，使图像居中
    rot_mat[0, 2] += (new_w / 2) - center[0]
    rot_mat[1, 2] += (new_h / 2) - center[1]

    # 执行旋转，指定新尺寸
    rotated = cv2.warpAffine(image, rot_mat, (new_w, new_h), flags=cv2.INTER_LINEAR)
    return rotated


def apply_scale_then_jpeg(image, scale_ratio, jpeg_quality, output_path):
    # 第一步：缩放
    h, w = image.shape[:2]
    resized = cv2.resize(image, (int(w * scale_ratio), int(h * scale_ratio)), interpolation=cv2.INTER_LINEAR)
    resized_back = cv2.resize(resized, (w, h), interpolation=cv2.INTER_LINEAR)

    # 第二步：JPEG压缩
    success = cv2.imwrite(output_path, resized_back, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
    if success:
        return cv2.imread(output_path)
    else:
        return resized_back  # 如果保存失败，返回未压缩的缩放图像

def apply_noise_then_jpeg(image, sigma, jpeg_quality, output_path):
    noisy_img = add_gaussian_noise(image, sigma)
    success = cv2.imwrite(output_path, noisy_img, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
    return cv2.imread(output_path) if success else noisy_img

def apply_MB_then_jpeg(image, sigma, jpeg_quality, output_path):
    MB_img = apply_mean_filter(image, sigma)
    success = cv2.imwrite(output_path, MB_img, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
    return cv2.imread(output_path) if success else MB_img

def apply_scale_then_noise(image, scale_ratio, sigma):
    h, w = image.shape[:2]
    resized = cv2.resize(image, (int(w * scale_ratio), int(h * scale_ratio)), interpolation=cv2.INTER_LINEAR)
    resized_back = cv2.resize(resized, (w, h), interpolation=cv2.INTER_LINEAR)

    return add_gaussian_noise(resized_back, sigma)


def apply_scale_then_noise_then_jpeg(image, scale_ratio, sigma,jpeg_quality, output_path):
    h, w = image.shape[:2]
    resized = cv2.resize(image, (int(w * scale_ratio), int(h * scale_ratio)), interpolation=cv2.INTER_LINEAR)
    resized_back = cv2.resize(resized, (w, h), interpolation=cv2.INTER_LINEAR)
    noisy_img = add_gaussian_noise(resized_back, sigma)
    success = cv2.imwrite(output_path, noisy_img, [cv2.IMWRITE_JPEG_QUALITY, jpeg_quality])
    return cv2.imread(output_path) if success else noisy_img


def add_gaussian_noise(image, sigma):
    image = image.astype(np.float32)
    noise = np.random.normal(0, sigma * 255, image.shape)
    noisy_image = image + noise
    noisy_image = np.clip(noisy_image, 0, 255).astype(np.uint8)
    return noisy_image

def add_salt_and_pepper_noise(image, amount):
    noisy_image = image.copy()
    num_pixels = image.size // image.shape[-1]  # Number of pixels
    num_salt = int(num_pixels * amount / 2)
    num_pepper = int(num_pixels * amount / 2)

    # Add salt (white) noise
    coords = [np.random.randint(0, i - 1, num_salt) for i in image.shape[:2]]
    noisy_image[coords[0], coords[1]] = 255

    # Add pepper (black) noise
    coords = [np.random.randint(0, i - 1, num_pepper) for i in image.shape[:2]]
    noisy_image[coords[0], coords[1]] = 0

    return noisy_image


def jpeg_compress(image, quality):
    encode_param = [int(cv2.IMWRITE_JPEG_QUALITY), quality]
    result, encimg = cv2.imencode('.jpg', image, encode_param)
    return cv2.imdecode(encimg, 1)

def scale_image(image, scale_ratio):
    h, w = image.shape[:2]
    resized = cv2.resize(image, (int(w * scale_ratio), int(h * scale_ratio)))
    return cv2.resize(resized, (w, h))


def extract_rotation_angle_from_affine(matrix):
    # 从仿射矩阵中提取旋转角度（单位：度）
    a, b = matrix[0, 0], matrix[0, 1]
    angle_rad = np.arctan2(b, a)
    angle_deg = np.rad2deg(angle_rad)
    return angle_deg


def center_crop(image, target_h, target_w):
    h_img, w_img = image.shape[:2]

    if h_img < target_h or w_img < target_w:
        raise ValueError("输入图像尺寸小于裁剪目标尺寸")

    top = (h_img - target_h) // 2
    left = (w_img - target_w) // 2

    return image[top:top + target_h, left:left + target_w]


def correct_image_rotation(img_ref, img_rotated):
    # 1. SIFT 特征提取
    sift = cv2.SIFT_create()
    kp1, des1 = sift.detectAndCompute(img_ref, None)
    kp2, des2 = sift.detectAndCompute(img_rotated, None)

    # 2. 特征匹配
    bf = cv2.BFMatcher()
    matches = bf.knnMatch(des1, des2, k=2)

    # 3. Lowe’s ratio test
    good_matches = []
    for m, n in matches:
        if m.distance < 0.75 * n.distance:
            good_matches.append(m)

    if len(good_matches) < 4:
        raise ValueError("匹配点太少，无法估计旋转")

    # 4. 提取匹配点
    pts1 = np.float32([kp1[m.queryIdx].pt for m in good_matches])
    pts2 = np.float32([kp2[m.trainIdx].pt for m in good_matches])

    # 5. 估计仿射矩阵
    M, inliers = cv2.estimateAffinePartial2D(pts2, pts1, method=cv2.RANSAC)
    if M is None:
        raise ValueError("无法估计仿射变换")

    # 6. 提取旋转角度
    angle = extract_rotation_angle_from_affine(M)
    print(f"估计旋转角度: {angle:.2f}°")
    angle= round(angle / 5) * 5

    # 7. 使用与你的 rotate_image 一致的方式反向旋转（完整画布 + 平移校正）
    (h_r, w_r) = img_rotated.shape[:2]
    center = (w_r / 2, h_r / 2)

    rot_mat = cv2.getRotationMatrix2D(center, angle, 1.0)
    abs_cos = abs(rot_mat[0, 0])
    abs_sin = abs(rot_mat[0, 1])
    new_w = int(h_r * abs_sin + w_r * abs_cos)
    new_h = int(h_r * abs_cos + w_r * abs_sin)

    rot_mat[0, 2] += (new_w / 2) - center[0]
    rot_mat[1, 2] += (new_h / 2) - center[1]

    rotated_full = cv2.warpAffine(img_rotated, rot_mat, (new_w, new_h), flags=cv2.INTER_LINEAR)
    (h_ref, w_ref) = img_ref.shape[:2]
    rotated_full= center_crop(rotated_full, h_ref, w_ref)

    output_path = f"{output_folder}/{image_name}_rotate_{angle}_image.bmp"
    cv2.imwrite(output_path, rotated_full)

    return rotated_full, angle




def get_common_corners(*corner_lists, threshold=1.5):
    base = corner_lists[0]
    common = []
    for pt in base:
        if all(np.min(np.linalg.norm(cl - pt, axis=1)) < threshold for cl in corner_lists[1:]):
            common.append(pt)
    return np.array(common, dtype=np.float32)

def triangle_area(a, b, c):
    return 0.5 * abs(np.cross(b - a, c - a))

def compute_angles(a, b, c):
    def angle(u, v):
        cos_val = np.dot(u, v) / (np.linalg.norm(u) * np.linalg.norm(v) + 1e-8)
        return degrees(np.arccos(np.clip(cos_val, -1.0, 1.0)))
    ab, bc, ca = b - a, c - b, a - c
    return [angle(-ab, ca), angle(-bc, ab), angle(-ca, bc)]

def remove_low_contribution_points(points, area_threshold=80.0, min_angle_deg=15, min_dist=18, max_iter=10):
    pts = points.copy()

    # 计算凸包，得到边界点索引
    hull = ConvexHull(pts)
    boundary_indices = set(hull.vertices)

    for iteration in range(max_iter):
        tri = Delaunay(pts)
        triangles = tri.simplices

        point_density = np.zeros(len(pts), dtype=int)
        for i, p in enumerate(pts):
            dists = np.linalg.norm(pts - p, axis=1)
            point_density[i] = np.sum(dists < min_dist) - 1

        points_to_remove = set()

        # 剔除密集点：但排除边界点
        for i, dens in enumerate(point_density):
            if i in boundary_indices:
                continue  # 边界点不剔除
            if dens >= 2:
                points_to_remove.add(i)

        # 检测差三角形，找三角形中密集度最高非边界点加入剔除集合
        for tri_idxs in triangles:
            a, b, c = pts[tri_idxs]
            area = triangle_area(a, b, c)
            angles = compute_angles(a, b, c)

            if area < area_threshold or min(angles) < min_angle_deg:
                # 在三角形点中排除边界点，选择密度最高点剔除
                non_boundary_pts = [idx for idx in tri_idxs if idx not in boundary_indices]
                if non_boundary_pts:
                    densities = [point_density[idx] for idx in non_boundary_pts]
                    max_density_idx = non_boundary_pts[np.argmax(densities)]
                    points_to_remove.add(max_density_idx)
                # 若三角形所有点均为边界点，则不剔除

        if not points_to_remove:
            break

        keep_mask = np.ones(len(pts), dtype=bool)
        for idx in points_to_remove:
            keep_mask[idx] = False

        pts = pts[keep_mask]

        # 重新计算边界点索引，保证边界保护持续有效
        hull = ConvexHull(pts)
        boundary_indices = set(hull.vertices)

    tri = Delaunay(pts)
    return pts, tri.simplices


def get_midpoints_from_corners(corners, image_shape, dis, min_dist):
    center = np.array([image_shape[1] / 2, image_shape[0] / 2])
    corners = sorted(corners, key=lambda pt: np.linalg.norm(pt - center))
    corners = np.array(corners)

    tri = Delaunay(corners)
    triangles = tri.simplices  # 索引数组

    # 绘制筛选后的基准点
    # 可视化绘制三角网格无中点
    result_img = np.ones_like(image) * 255
    # 绘制角点
    for pt in corners:
        cv2.circle(result_img, tuple(np.round(pt).astype(int)), 4, (114, 83, 51), -1)
    output_path = f"{output_folder}/{image_name}_Just_midpoint_before_image.bmp"
    cv2.imwrite(output_path, result_img)

    # 绘制初始网格
    # 可视化绘制三角网格无中点
    result_img = np.ones_like(image) * 255
    for triangle in triangles:
        pts = corners[triangle].reshape(-1, 1, 2).astype(np.int32)
        cv2.polylines(result_img, [pts], True, (114, 83, 51), 2)
    # 绘制角点
    for pt in corners:
        cv2.circle(result_img, tuple(np.round(pt).astype(int)), 4, (114, 83, 51), -1)
    output_path = f"{output_folder}/{image_name}_midpoint_before_image.bmp"
    cv2.imwrite(output_path, result_img)

    # 网格简化
    corners, triangles = remove_low_contribution_points(corners)

    # 绘制简化网格
    # 可视化绘制三角网格无中点
    result_img = np.ones_like(image) * 255
    for triangle in triangles:
        pts = corners[triangle].reshape(-1, 1, 2).astype(np.int32)
        cv2.polylines(result_img, [pts], True, (114, 83, 51), 2)
    # 绘制角点
    for pt in corners:
        cv2.circle(result_img, tuple(np.round(pt).astype(int)), 4, (114, 83, 51), -1)
    output_path = f"{output_folder}/{image_name}_midpoint_after_image.bmp"
    cv2.imwrite(output_path, result_img)

    midpoints = []

    def should_add(pt, pts):
        return all(np.linalg.norm(pt - p) >= min_dist for p in pts)

    for simplex in triangles:
        triangle = corners[simplex]  # 获取三角形顶点坐标

        edge_divisions = []

        # Step 1：边上等分点
        for i in range(3):
            pt1, pt2 = triangle[i], triangle[(i + 1) % 3]
            d = np.linalg.norm(pt1 - pt2)
            num_parts = int(d // dis )
            edge_pts = []
            if num_parts >= 1:
                for k in range(1, num_parts + 1):
                    new_pt = pt1 + (pt2 - pt1) * (k / (num_parts + 1))
                    if should_add(new_pt, midpoints):
                        midpoints.append(new_pt)
                    edge_pts.append(new_pt)
            edge_divisions.append(edge_pts)

        # Step 2：内部边分点连线采样
        for i in range(3):
            edge1 = edge_divisions[i]
            edge2 = edge_divisions[(i + 1) % 3]
            for pt1 in edge1:
                for pt2 in edge2:
                    line_vec = pt2 - pt1
                    line_len = np.linalg.norm(line_vec)
                    num_internal_parts = int(line_len // dis )
                    if num_internal_parts >= 1:
                        for j in range(1, num_internal_parts + 1):
                            internal_pt = pt1 + line_vec * (j / (num_internal_parts + 1))
                            if should_add(internal_pt, midpoints):
                                midpoints.append(internal_pt)

    # 可视化绘制三角网格无中点
    result_img = image.copy()
    for triangle in triangles:
        pts = corners[triangle].reshape(-1, 1, 2).astype(np.int32)
        cv2.polylines(result_img, [pts], True, (255, 255, 0), 1)

    # 绘制角点
    for pt in corners:
        cv2.circle(result_img, tuple(np.round(pt).astype(int)), 4, (255, 255, 0), -1)

    # 绘制中点
    for pt in midpoints:
        center = tuple(np.round(pt).astype(int))
        d = 4
        pts = np.array([
            [center[0], center[1] - d],
            [center[0] + d, center[1]],
            [center[0], center[1] + d],
            [center[0] - d, center[1]]
        ], np.int32).reshape((-1, 1, 2))
        cv2.fillPoly(result_img, [pts], (0, 255, 255))

    output_path = f"{output_folder}/{image_name}_midpoint_image.bmp"
    cv2.imwrite(output_path, result_img)

    # 可视化绘制三角网格无中点
    result_img = np.ones_like(image) * 255
    for triangle in triangles:
        pts = corners[triangle].reshape(-1, 1, 2).astype(np.int32)
        cv2.polylines(result_img, [pts], True, (114, 83, 51), 2)

    # 绘制角点
    for pt in corners:
        cv2.circle(result_img, tuple(np.round(pt).astype(int)), 4, (114, 83, 51), -1)

    # 绘制中点
    for pt in midpoints:
        center = tuple(np.round(pt).astype(int))
        d = 4
        pts = np.array([
            [center[0], center[1] - d],
            [center[0] + d, center[1]],
            [center[0], center[1] + d],
            [center[0] - d, center[1]]
        ], np.int32).reshape((-1, 1, 2))
        cv2.fillPoly(result_img, [pts], (89, 86, 226))

    output_path = f"{output_folder}/{image_name}_midpoint_no_image.bmp"
    cv2.imwrite(output_path, result_img)

    return np.array(midpoints)



def sort_midpoints_by_center(midpoints, image_shape):
    center = np.array([image_shape[1] / 2, image_shape[0] / 2])
    distances = [np.linalg.norm(pt - center) for pt in midpoints]
    sorted_indices = np.argsort(distances)
    sorted_midpoints = midpoints[sorted_indices]
    return sorted_midpoints

def visualize_sorted_midpoints(image, sorted_midpoints, output_path=None):
    vis_img = image.copy()
    font = cv2.FONT_HERSHEY_SIMPLEX

    for idx, pt in enumerate(sorted_midpoints):
        pt_int = tuple(np.round(pt).astype(int))
        cv2.circle(vis_img, pt_int, 3, (0, 0, 255), -1)  # 红点
        cv2.putText(vis_img, str(idx), (pt_int[0] + 3, pt_int[1] - 3),
                    font, 0.4, (0, 255, 0), 1, cv2.LINE_AA)

    # if output_path:
    cv2.imwrite(output_path, vis_img)
    # else:
    #     cv2.imshow("Sorted Midpoints", vis_img)
    #     cv2.waitKey(0)
    #     cv2.destroyAllWindows()

def crop_region(img):
    img_copy = np.copy(img)  # 复制原图，避免修改原始图像
    h, w = img.shape[:2]
    img_copy[50:150, w-250:w-50] = (0, 0, 0)  # 裁剪右上角区域为黑色
    return img_copy

def binary_to_quaternary(binary_str):
    # 补足位数到4的倍数
    padding = (2 - len(binary_str) % 2) % 2
    # binary_str = '0' * padding + binary_str
    binary_str = binary_str + '0' * padding

    quaternary_str = ''
    for i in range(0, len(binary_str), 2):
        bits = binary_str[i:i + 2]
        quaternary_digit = str(int(bits, 2))  # 每两个bit转成一个四进制数
        quaternary_str += quaternary_digit
    return quaternary_str


def quaternary_to_binary(quaternary_str):
    binary_str = ''
    for digit in quaternary_str:
        binary_bits = format(int(digit), '02b')  # 每个四进制位转为2位二进制
        binary_str += binary_bits
    return binary_str


def rgb_to_ycbcr_float(rgb_block):
    rgb = rgb_block.astype(np.float32)
    R = rgb[:, :, 2]
    G = rgb[:, :, 1]
    B = rgb[:, :, 0]

    Y  =0.299 * R + 0.587 * G + 0.114 * B
    Cb = -0.168736 * R - 0.331264 * G + 0.5 * B + 128
    Cr = 0.5 * R - 0.418688 * G - 0.081312 * B + 128

    ycbcr = np.stack((Y, Cb, Cr), axis=-1)
    return ycbcr  # float32 YCbCr

def ycbcr_to_rgb_float(ycbcr_block):
    # _normalized
    Y  = ycbcr_block[:, :, 0]
    Cb = ycbcr_block[:, :, 1]
    Cr = ycbcr_block[:, :, 2]

    R = Y + 1.402 * (Cr - 128)
    G = Y - 0.344136 * (Cb - 128) - 0.714136 * (Cr - 128)
    B = Y + 1.772 * (Cb - 128)

    rgb = np.stack((B, G, R), axis=-1)

    # 计算每个像素的最大值
    max_rgb = np.max(rgb, axis=2, keepdims=True)

    # 对超过255的像素点，按比例缩放
    scale = np.ones_like(max_rgb)
    over_mask = max_rgb > 255
    scale[over_mask] = 255.0 / max_rgb[over_mask]

    rgb_scaled = rgb * scale

    return np.clip(rgb_scaled, 0, 255)  # float32 RGB


def modify_block_by_mean(image_s, old_vals, new_vals, x, y, mean_size_1, mean_size_2):
    block_size = 2 * mean_size_1 + mean_size_2
    half_block = block_size // 2


    top = y - half_block
    left = x - half_block
    bottom = top + block_size
    right = left + block_size

    block_rgb = image_s[top:bottom, left:right]
    block_ycbcr = rgb_to_ycbcr_float(block_rgb)

    Y = block_ycbcr[:, :, Channel].astype(np.float32)

    sub_size = mean_size_1
    coords = {
        'A': (0, 0),
        'B': (0, block_size - sub_size),
        'C': (block_size - sub_size, 0),
        'D': (block_size - sub_size, block_size - sub_size),
    }
    regions = ['A', 'B', 'C', 'D']


    for idx, region in enumerate(regions):
        old, new = old_vals[idx], new_vals[idx]
        if old == new:
            continue

        best_change = new - old
        total_change = (new - old) * ((half_block + 1) ** 2)
        # print(f"new - old为{new - old}，total_change为{total_change}")

        dy, dx = coords[region]
        y1, y2 = dy, dy + sub_size
        x1, x2 = dx, dx + sub_size

        sub_block = Y[y1:y2, x1:x2].copy()
        original_block = sub_block.copy()
        # print(f"初始 this is {idx}块")
        # print(f"{original_block}")

        h, w = sub_block.shape

        flat_vals = sub_block.flatten()
        diffs = np.abs(flat_vals[:, None] - flat_vals[None, :])
        close_pairs = np.sum(diffs < 10)
        total_pairs = diffs.size

        # if embed_index != 47:
        if embed_index < 0:
            continue
        else:
            # grad = cv2.Laplacian(sub_block, cv2.CV_32F)
            # grad = np.abs(grad)
            grad = cv2.Sobel(sub_block, cv2.CV_32F, 1, 1, ksize=3)
            grad = np.abs(grad)
            norm_grad = grad / (np.max(grad) + 1e-6)

            cy, cx = h / 2, w / 2
            Y_idx, X_idx = np.indices((h, w))
            dist = np.sqrt((Y_idx - cy) ** 2 + (X_idx - cx) ** 2)
            max_dist = np.max(dist)
            # 这里要使距离中心越远（越靠近边缘）的像素，修改权值越小，所以应该1-（）
            # norm_dist =  (dist / (max_dist + 1e-6))
            norm_dist = 1 - (dist / (max_dist + 1e-6))

            # norm_grad = grad / (np.max(grad) + 1e-6)
            alpha = 0.7
            weights = alpha * norm_grad + (1 - alpha) * norm_dist

            # if total_change > 0:
            #     # 正向修改：权重 ∝ 边缘强度（grad大 → 分配多）
            #     weights = alpha * norm_grad + (1 - alpha) * norm_dist
            # else:
            #     # 负向修改：权重 ∝ 1 / grad（grad大 → 分配少）
            #     inverse_grad = 1.0 - norm_grad  # 等价于“边缘弱 → 分配多”
            #     weights = alpha * inverse_grad + (1 - alpha) * norm_dist

            weights_sum = np.sum(weights) + 1e-6
            weights /= weights_sum

            for attempt in range(5):  # 最多尝试3轮
                delta_map = weights * total_change
                for i in range(h):
                    for j in range(w):
                        val = sub_block[i, j]
                        delta = delta_map[i, j]
                        # new_val = np.clip(val + delta, 20, 156)
                        new_val = np.clip(val + delta, 1, 253)
                        actual_delta = new_val - val
                        sub_block[i, j] = new_val
                        total_change -= actual_delta
                if abs(total_change) <= 4:
                    break

            if abs(total_change) > 6:
                print(f"一轮total change: {total_change}")
                avg_change = total_change / (h * w)
                for i in range(h):
                    for j in range(w):
                        val = sub_block[i, j]
                        # new_val = np.clip(val + avg_change, 20, 156)
                        new_val = np.clip(val + avg_change, 1, 253)
                        actual_delta = new_val - val
                        sub_block[i, j] = new_val
                        total_change -= actual_delta

            if abs(total_change) > 6:
                print(f"二轮total change: {total_change}")
                for attempt in range(5):  # 最多尝试3轮
                    if abs(total_change) <= 6:
                        break
                    adjust_value = 5 if total_change > 0 else -5
                    for i in range(h):
                        if abs(total_change) <= 5:
                            break
                        for j in range(w):
                            if abs(total_change) <= 5:
                                break
                            val = sub_block[i, j]
                            new_val = val + adjust_value
                            # if 2 <= new_val <= 254:
                            if 1 <= new_val <= 253:
                                sub_block[i, j] = new_val
                                total_change -= adjust_value
                # print ()
                # 输出像素值
            if abs(total_change) > 4:
                print(f"【{region}块】像素修改前后（float精度）:")
                print("原始像素:")
                print(np.round(original_block, 2))
                print("修改后像素:")
                print(np.round(sub_block, 2))



        # Y[y1:y2, x1:x2] = np.clip(sub_block, 20, 156).astype(np.float32)
        # print(f"修改后this is {idx}块")
        # print(f"{sub_block}")
        Y[y1:y2, x1:x2] = np.clip(sub_block, 1, 253).astype(np.float32)

    # block_ycbcr[:, :, Channel] = Y.clip(20, 156).astype(np.float32)
    block_ycbcr[:, :, Channel] = Y.clip(1, 253).astype(np.float32)
    # print(f"修改后this is Y通道 ")
    # print(f"{block_ycbcr[:, :, Channel]}")
    # print(f"修改后this is Cb通道")
    # print(f"{block_ycbcr[:, :, 1]}")
    # print(f"修改后this is Cr通道")
    # print(f"{block_ycbcr[:, :, 2]}")

    image_s[top:bottom, left:right] = ycbcr_to_rgb_float(block_ycbcr)


def classify_matrix(M, energy_threshold):
    A, B = M[0]
    C, D = M[1]

    AB = B - A
    CD = D - C
    AC = C - A
    BD = D - B

    diffs = {'AB': AB, 'CD': CD, 'AC': AC, 'BD': BD}
    abs_sorted = sorted(diffs.items(), key=lambda x: abs(x[1]), reverse=True)
    top2 = abs_sorted[:2]
    low2 = abs_sorted[2:]

    top_keys = [k for k, _ in top2]
    top_values = [v for _, v in top2]
    low_values = [v for _, v in low2]
    max_diff = max(abs(v) for v in diffs.values())
    min_diff = min(abs(v) for v in diffs.values())
    all_values = list(diffs.values())

    total_abs_sum = sum(abs(v) for v in all_values)

    if  all(abs(top) - abs(low) >= energy_threshold for top in top_values for low in low_values)  :
        # 原有Type 1~3判断
        if ('AB' in top_keys and 'CD' in top_keys) or ('AC' in top_keys and 'BD' in top_keys):
            if all(v > 0 for v in top_values) or all(v < 0 for v in top_values):
                return 1

        # Type 2/3: 高梯度交叉，起点相同 or 终点相同
        edges = {
            'AB': ('A', 'B') if AB > 0 else ('B', 'A'),
            'CD': ('C', 'D') if CD > 0 else ('D', 'C'),
            'AC': ('A', 'C') if AC > 0 else ('C', 'A'),
            'BD': ('B', 'D') if BD > 0 else ('D', 'B'),
        }

        if all(k in edges for k in top_keys):
            start1, end1 = edges[top_keys[0]]
            start2, end2 = edges[top_keys[1]]
            if start1 == start2:
                return 2  # 起点相同
            elif end1 == end2:
                return 3  # 终点相同



    # Type 6: 所有边差值都小，但略大于type5
    if  all(abs(v) < energy_threshold/4 for v in all_values) or all(abs(top) - abs(low) <= 1 for top in top_values for low in low_values):
        return 0


    return -1


def classify_matrix_extract(M, energy_threshold):
    A, B = M[0]
    C, D = M[1]


    AB = B - A
    CD = D - C
    AC = C - A
    BD = D - B

    diffs = {'AB': AB, 'CD': CD, 'AC': AC, 'BD': BD}
    abs_sorted = sorted(diffs.items(), key=lambda x: abs(x[1]), reverse=True)
    top2 = abs_sorted[:2]
    low2 = abs_sorted[2:]

    top_keys = [k for k, _ in top2]
    top_values = [v for _, v in top2]
    # low_values = [v for _, v in low2]
    max_diff = max(abs(v) for v in diffs.values())
    min_diff = min(abs(v) for v in diffs.values())

    # 提取原始A、B、C、D数值
    value_list = [A, B, C, D]
    sorted_values = sorted(value_list)  # 从小到大排序


    if max_diff - min_diff < energy_threshold / 2 +2 and abs(
            abs(abs_sorted[1][1]) - abs(abs_sorted[2][1])) < energy_threshold / 2:
        return 0
    # if 2>1:
    else:
        if ('AB' in top_keys and 'CD' in top_keys) or ('AC' in top_keys and 'BD' in top_keys):
            if all(v > 0 for v in top_values) or all(v < 0 for v in top_values):
                return 1

        # Type 2/3: 高梯度交叉，起点相同 or 终点相同
        edges = {
            'AB': ('A', 'B') if AB > 0 else ('B', 'A'),
            'CD': ('C', 'D') if CD > 0 else ('D', 'C'),
            'AC': ('A', 'C') if AC > 0 else ('C', 'A'),
            'BD': ('B', 'D') if BD > 0 else ('D', 'B'),
        }

        if all(k in edges for k in top_keys):
            start1, end1 = edges[top_keys[0]]
            start2, end2 = edges[top_keys[1]]
            if start1 == start2:
                return 2  # 起点相同
            elif end1 == end2:
                return 3  # 终点相同
    return 0




def get_string_form_watermark(watermark):
    # 初始化一个空的字符串来保存结果
    binary_string = ""
    # 遍历图像的每个像素，将黑色和白色转换为二进制字符
    for row in water_i_gray:
        for pixel in row:
            # 如果像素值是255（白色），则添加1；如果是0（黑色），则添加0
            if pixel >= 200:
                binary_string += "1"
            elif pixel <= 100:
                binary_string += "0"
    return binary_string

def get_midpoints_from_image(image):
    # 图像攻击

    image_jpeg = jpeg_compress(image, quality=80)
    image_scaled = scale_image(image, scale_ratio=0.8)
    # image_noise = add_gaussian_noise(image, sigma=0.01)

    max_corners = 400

    # SIFT角点检测
    corners_orig = sift_corner_detection(image, max_corners)
    corners_jpeg = sift_corner_detection(image_jpeg, max_corners)
    corners_scaled = sift_corner_detection(image_scaled, max_corners)
    # corners_noise = sift_corner_detection(image_noise, max_corners)

    # 获取共同角点
    common_corners = get_common_corners(corners_orig,  corners_jpeg, corners_scaled)

    # 获取定位点
    midpoints = get_midpoints_from_corners(common_corners, image.shape[:2],midpoint_distance,midpoint_distance)
    print(f"一轮筛选后len(midpoints)为{len(midpoints)}")

    # 计算block_radius
    block_radius = int((2 * mean_size_1 + mean_size_2) / 2)
    block_radius = int(round(block_radius))
    h, w, _ = image.shape

    # 边缘判断：过滤距离图像边缘太近的点
    filtered_midpoints = []
    for pt in midpoints:
        x, y = int(round(pt[0])), int(round(pt[1]))
        if block_radius < x < w - block_radius-1 and block_radius < y < h - block_radius-1:
            filtered_midpoints.append((x, y))
    midpoints = np.array(filtered_midpoints)
    print(f"边缘筛选后len(midpoints)为{len(midpoints)}")

    # 2. 中心排序
    midpoints = sort_midpoints_by_center(midpoints, image.shape[:2])
    return midpoints


def generate_prioritized_deltas(max_abs_delta):
    # 生成以 0 为中心，变化量逐渐增大的顺序
    deltas = [0]
    for i in range(1, max_abs_delta + 1):
        deltas.extend([i, -i])
    return deltas


def search_modifications(target_class, num_values_to_change, threshold, range01, range02):
    from itertools import combinations, product

    all_indices = [0, 1, 2, 3]  # 对应 A B C D
    best_new_vals = None
    min_adjusted_cost = float('inf')


    for idxs in combinations(all_indices, num_values_to_change):
        delta_ranges = []
        for i in idxs:
            val = int(eval(f"{chr(65 + i)}"))  # 获取 A/B/C/D 值
            if val < 22:
                delta_range = [d for d in range(0, range01) if 8 <= val + d <= 250]
            elif val < 40:
                delta_range = [d for d in range(-range01, range01) if 20 <= val + d <= 250]
            elif val > 80:
                delta_range = [d for d in range(-range02, range02) if 8 <= val + d <= 250]
            else:
                delta = int(abs(val) * M_Percentage)
                # delta_range = delta_lookup[delta]
                delta_range = generate_prioritized_deltas(delta)
            delta_ranges.append(delta_range)

        max_trials = 20000000
        trial_count = 0
        for delta_vals in product(*delta_ranges):
            trial_count += 1
            if trial_count > max_trials:
                print(f"超过最大尝试次数，停止搜索，x,y={x},{y}, 目标类别={target_class}")
                return best_new_vals

            new_vals = [A, B, C, D]
            skip = False
            for i, delta in zip(idxs, delta_vals):
                new_val = new_vals[i] + delta
                if new_val < 8 or new_val > 250:
                    skip = True
                    break
                new_vals[i] = new_val
            if skip:
                continue

            new_M = np.array([[new_vals[0], new_vals[1]], [new_vals[2], new_vals[3]]])
            result = classify_matrix(new_M, threshold)

            if result == target_class:
                # 用局部MSE替代原始修改幅度
                old_vals = [A, B, C, D]
                mse_cost = np.mean([(new_vals[i] - old_vals[i]) ** 2 for i in idxs])

                adjusted_cost = mse_cost

                if adjusted_cost < min_adjusted_cost:
                    best_new_vals = new_vals
                    min_adjusted_cost = adjusted_cost
                    if adjusted_cost < 0.5:
                        return best_new_vals

    return best_new_vals


start_row = 1

for config_index, (mean_size_1, watermark_shape, midpoint_distance, watermark_file) in enumerate(config_list):

    # 写入 header（包括攻击方式下的 NC 和 BER）
    header = [f"{mean_size_1}", "PSNR", "SSIM"]
    for mode_name, mode_function in modes:
        header.extend([f"{mode_name}(NC)", f" "])
    ws.append(header)

    mean_size_2 = 1
    watermark_length = watermark_shape * watermark_shape
    Channel = 0

    # 读取水印
    water_i = cv2.imread(watermark_file)
    water_i_gray = cv2.cvtColor(water_i, cv2.COLOR_BGR2GRAY)

    print(f"\n====== 当前参数组 {config_index+1}：mean_size_1={mean_size_1}, watermark_shape={watermark_shape}, midpoint_distance={midpoint_distance} ======\n")

    # 批量处理
    for image_path in image_list:

        image = cv2.imread(image_path)
        image_ini = image.copy()
        image_name = os.path.splitext(os.path.basename(image_path))[0]

        output_folder = f"__Try__{mean_size_1}__ene__{energy_threshold}_{image_name}___"
        os.makedirs(output_folder, exist_ok=True)

        print(f"Processing {image_name}...")

        midpoints = get_midpoints_from_image(image)

        midpoints_stego = []
        # 3. 可视化结果
        visualize_sorted_midpoints(image, midpoints,
                                   output_path=f"{output_folder}/{image_name}_sorted_midpoints.bmp")

        X1 = get_string_form_watermark(water_i_gray)
        # 得到四进制数据
        X2 = binary_to_quaternary(X1)
        num_X2 = len(X2)

        block_radius = int((2 * mean_size_1 + mean_size_2) / 2)
        block_radius = int(round(block_radius))

        # 创建日志文件并重定向 stdout
        output_path = f"{output_folder}/{image_name}_output_log.txt"
        log_file = open(output_path, "w", encoding="utf-8")
        sys.stdout = log_file

        print(
            f"\n====== 当前参数组 {config_index + 1}：mean_size_1={mean_size_1}, watermark_shape={watermark_shape}, midpoint_distance={midpoint_distance} ======\n")

        print(
            f"mean_size 1 is:{mean_size_1},mean_size 2 is:{mean_size_2},watermark_shape：{watermark_shape}，midpoint_distance：{midpoint_distance},energy_threshold :{energy_threshold}")

        print(f"嵌入点数量为{len(midpoints)}")

        image_s = image.copy()  # 用于保存修改后图像

        # 设定参数
        embed_index = 0
        wrong_test = []
        one_not_enough = 0
        two_not_enough = 0
        four_not_enough = 0
        six_not_enough = 0
        eight_not_enough = 0
        can_not_find_modifi = 0
        type5 = 0
        type6 = 0
        type4 = 0
        from6to5_false = 0
        h, w, _ = image.shape

        M03 = 0

        global_min = float('inf')
        global_max = float('-inf')

        # 进行数据隐藏
        for (x, y) in midpoints:
            if embed_index >= num_X2:
                break

            # if embed_index!=46:
            #     embed_index+=1
            #     continue

            x = int(round(x))
            y = int(round(y))

            if x - block_radius < 0 or y - block_radius < 0 or x + block_radius + 1 >= w or y + block_radius + 1 >= h:
                continue

            Q = image_s[y - block_radius:y + block_radius + 1, x - block_radius:x + block_radius + 1]
            Q_ycrcb = rgb_to_ycbcr_float(Q)
            Y_block = Q_ycrcb[:, :, Channel].astype(np.float32)

            # print(f"初始值")
            # print(f"{Y_block}")

            # 更新全局最小值和最大值
            local_min = Y_block.min()
            local_max = Y_block.max()
            if local_min < global_min:
                global_min = local_min
            if local_max > global_max:
                global_max = local_max

            size = block_radius + 1
            A = np.mean(Y_block[0:size, 0:size])
            B = np.mean(Y_block[0:size, -size:])
            C = np.mean(Y_block[-size:, 0:size])
            D = np.mean(Y_block[-size:, -size:])
            M = np.array([[A, B], [C, D]])
            current_type = classify_matrix(M, energy_threshold)

            digit = int(X2[embed_index])

            # print(f"index为{embed_index}，初始x,y为{x, y},vals为{A, B, C, D}，")
            print(f"index为{embed_index}，——————————————————————————")
            print(f"初始x,y为{x, y},vals为A={A:.2f},B={B:.2f}, C={C:.2f},D={D:.2f}，,待嵌入digit为{digit} ")

            M_max, M_min = np.max(M), np.min(M)
            M_range = M_max - M_min
            M_Percentage = M_range / (M_min + 1e-5)

            if M_Percentage > 0.3:
                M03 += 1
                M_Percentage = 0.3

            # embed_index += 1

            all_indices = [0, 1, 2, 3]  # 对应 A B C D
            best_new_vals = None
            min_change_sum = float('inf')

            if digit == 0:
                new_vals = search_modifications(target_class=0, num_values_to_change=3, threshold=energy_threshold,
                                                range01=10, range02=15)
            else:
                new_vals = search_modifications(target_class=digit, num_values_to_change=2,
                                                threshold=energy_threshold,
                                                range01=10, range02=15)

            if new_vals is None:
                one_not_enough += 1
                new_vals = search_modifications(target_class=digit, num_values_to_change=3,
                                                threshold=energy_threshold,
                                                range01=10, range02=15)

            if new_vals is None:
                two_not_enough += 1
                # print(f"M_Percentage为{M_Percentage}")
                M_Percentage = 0.3

                if digit == 0:
                    new_vals = search_modifications(target_class=0, num_values_to_change=4,
                                                    threshold=energy_threshold,
                                                    range01=15, range02=20)
                else:
                    new_vals = search_modifications(target_class=digit, num_values_to_change=3,
                                                    threshold=energy_threshold, range01=15, range02=20)

                # continue  # 或者降value_range重试也可以

                if new_vals is None:
                    # four_not_enough += 1
                    M_Percentage = 0.5
                    new_vals = search_modifications(target_class=digit, num_values_to_change=4,
                                                    threshold=energy_threshold, range01=20, range02=25)

                    # print(f"M_Percentage为{M_Percentage}")

            # 检验
            size = block_radius + 1  # 保证与前面一致

            old_vals = [A, B, C, D]

            if new_vals is None:
                four_not_enough += 1
                M_Percentage = 0.5
                new_vals = search_modifications(target_class=digit, num_values_to_change=4,
                                                threshold=energy_threshold,
                                                range01=30, range02=30)
            if new_vals is None:
                six_not_enough += 1
                M_Percentage = 3
                new_vals = search_modifications(target_class=digit, num_values_to_change=4,
                                                threshold=energy_threshold,
                                                range01=80, range02=80)

            if new_vals is None:
                eight_not_enough += 1
                print("new_val is none")

                print(
                    f"目前的x,y为{x, y},vals为A={A:.2f}, B={B:.2f}, C={C:.2f}, D={D:.2f}，digit为{digit}")

            else:

                # 否则使用原本 new_vals
                modify_block_by_mean(image_s, old_vals, new_vals, x, y, mean_size_1, mean_size_2)

                print(
                    f"选择的x,y为{x, y},vals为A={new_vals[0]:.2f}, B={new_vals[1]:.2f}, C={new_vals[2]:.2f}, D={new_vals[3]:.2f}，digit为{digit}")
            midpoints_stego.append((x, y))
            wrong_test.append((x, y))
            embed_index += 1  # 只有在成功嵌入digit时才 +1

            Q = image_s[y - block_radius:y + block_radius + 1, x - block_radius:x + block_radius + 1]
            Q_ycrcb = rgb_to_ycbcr_float(Q)
            Y_block = Q_ycrcb[:, :, Channel].astype(np.float32)

            size = block_radius + 1
            A = np.mean(Y_block[0:size, 0:size])
            B = np.mean(Y_block[0:size, -size:])
            C = np.mean(Y_block[-size:, 0:size])
            D = np.mean(Y_block[-size:, -size:])
            # midpoint_means[(x, y)] = (A, B, C, D)
            print(f"实际上x,y为{x, y},vals为A={A:.2f}, B={B:.2f}, C={C:.2f}, D={D:.2f}，digit为{digit}")

            # print(f"最终this is Y通道 ")
            # print(f"{Q_ycrcb[:, :, Channel]}")
            # print(f"最终this is Cb通道")
            # print(f"{Q_ycrcb[:, :, 1]}")
            # print(f"最终this is Cr通道")
            # print(f"{Q_ycrcb[:, :, 2]}")



        # 6. 生成最终图像I2
        image_s = image_s

        # check_image_s_range(image_s)

        image_stego = image_s

        output_path = f"{output_folder}/{image_name}_stego_image.bmp"
        cv2.imwrite(output_path, image_stego)


        def compute_psnr(cover, stego):
            return psnr(cover, stego, data_range=255)


        def calculate_ssim(img1, img2):
            return ssim(img1, img2, channel_axis=-1, data_range=255)


        # 计算 PSNR
        psnr_image = compute_psnr(image_ini, image_stego)
        ssim_image = calculate_ssim(image_ini, image_stego)
        print(f"{image_name}_图像PSNR 值: {psnr_image:.2f} dB")
        print(f"{image_name}_图像SSIM 值: {ssim_image:.4f} ")


        row_data = [image_name, psnr_image, ssim_image]

        for mode_name, mode_function in modes:

            print(f"{mode_name} processing finished for {image_name}.")

            print(
                f"\n====== 当前参数组 {config_index + 1}：mean_size_1={mean_size_1}, watermark_shape={watermark_shape}, midpoint_distance={midpoint_distance} ======\n")

            print(
                f"mean_size 1 is:{mean_size_1},mean_size 2 is:{mean_size_2},watermark_shape：{watermark_shape}，midpoint_distance：{midpoint_distance},energy_threshold :{energy_threshold}")

            print(f"嵌入点数量为{len(midpoints)}")

            output_path = f"{output_folder}/{image_name}_stego_image.bmp"
            cv2.imwrite(output_path, image_stego)
            image_stego = cv2.imread(output_path)

            # 处理当前模式
            # 攻击隐写图像
            output_path = os.path.join(output_folder, f"{mode_name}_output.jpg")
            attacked_image = mode_function(image_stego, output_path)
            cv2.imwrite(output_path, attacked_image)

            # attacked_image, angle = correct_image_rotation(image, attacked_image)

            gray = cv2.cvtColor(image_s, cv2.COLOR_BGR2GRAY)

            # 提取
            image_c = image_s.copy()

            get_nothing_extract = 0

            # midpoints = get_midpoints_from_image(image)

            # 4. 提取秘密信息
            E4 = ''
            h, w, _ = image.shape
            extract_index = 0

            for (x, y) in midpoints:

                # 提取数据长度达到水印长度，则截断
                if len(E4) >= int((watermark_length + 1) / 2):
                    E4 = E4[:int((watermark_length + 1) / 2)]
                    break

                # if extract_index!=46:
                #     extract_index+=1
                #     continue

                x = int(round(x))
                y = int(round(y))

                if x - block_radius < 0 or y - block_radius < 0 or x + block_radius + 1 >= w or y + block_radius + 1 >= h:
                    continue

                Q = attacked_image[y - block_radius:y + block_radius + 1, x - block_radius:x + block_radius + 1]
                Q_ycrcb = rgb_to_ycbcr_float(Q)
                Y_block = Q_ycrcb[:, :, Channel].astype(np.float32)

                # print(f"提取时 this is 块")
                # print(f"{Y_block}")

                size = block_radius + 1
                A = np.mean(Y_block[0:size, 0:size])
                B = np.mean(Y_block[0:size, -size:])
                C = np.mean(Y_block[-size:, 0:size])
                D = np.mean(Y_block[-size:, -size:])
                M = np.array([[A, B], [C, D]])

                result = classify_matrix_extract(M, energy_threshold)
                if result in [0, 1, 2, 3]:
                    E4 += str(result)
                    print(
                        f"提取x,y为{x, y},index为{extract_index}，vals为A={A:.2f}, B={B:.2f}, C={C:.2f}, D={D:.2f}，digit为{result}")
                else:
                    E4 += str(0)
                    get_nothing_extract += 1
                    print(
                        f"get nothing but{result},提取x,y为{x, y},vals为A={A:.2f}, B={B:.2f}, C={C:.2f}, D={D:.2f}，digit为{result}")
                    # E4 += str(0)

                # 提取数据长度达到水印长度，则截断
                if len(E4) >= int((watermark_length + 1) / 2):
                    E4 = E4[:int((watermark_length + 1) / 2)]
                    break

                extract_index += 1

            E2 = quaternary_to_binary(E4)
            print(f"E2为{E2}")

            if len(E2) >= watermark_length:
                E2 = E2[:watermark_length]

            # 确保字符串长度为400
            print(f"提取的隐写字符串长度:{len(E2)}")
            print(f"X1为{X1}")
            print(f"E2为{E2}")
            print(f"X2为{X2}")
            print(f"E4为{E4}")


            # 确保两个字符串的长度相等
            if len(E2) != len(X1):
                print(f"X1长度为{len(X1)},{mode_name}_两个字符串长度不同，无法计算百分比。")
            else:
                # 计算不同字符的个数
                diff_count = sum(1 for i in range(len(E2)) if E2[i] != X1[i])

                # 计算不同字符的百分比
                diff_percentage = (diff_count / len(E2)) * 100

                print(f"{mode_name}_不同字符的百分比: {diff_percentage:.2f}%")

            wrong_index = 0
            for (x, y) in midpoints:
                # wrong_index+=1
                if wrong_index >= len(X2):
                    break
                if X2[wrong_index] != E4[wrong_index]:
                    x = int(round(x))
                    y = int(round(y))
                    block_radius = int((2 * mean_size_1 + mean_size_2) / 2)
                    block_radius = int(round(block_radius))
                    Q = image_stego[y - block_radius:y + block_radius + 1, x - block_radius:x + block_radius + 1]
                    Q_ycrcb = rgb_to_ycbcr_float(Q)
                    Y_block = Q_ycrcb[:, :, Channel].astype(np.float32)
                    size = block_radius + 1
                    A = np.mean(Y_block[0:size, 0:size])
                    B = np.mean(Y_block[0:size, -size:])
                    C = np.mean(Y_block[-size:, 0:size])
                    D = np.mean(Y_block[-size:, -size:])
                    digit = X2[wrong_index]
                    print(f"字符错误：索引 {wrong_index}，坐标 = ({x}, {y})")
                    print(f"隐写x,y为{x, y},vals为A={A:.2f}, B={B:.2f}, C={C:.2f}, D={D:.2f}，digit为{digit}")
                    x = int(round(x))
                    y = int(round(y))
                    block_radius = int((2 * mean_size_1 + mean_size_2) / 2)
                    block_radius = int(round(block_radius))
                    Q = attacked_image[y - block_radius:y + block_radius + 1, x - block_radius:x + block_radius + 1]
                    Q_ycrcb = rgb_to_ycbcr_float(Q)
                    Y_block = Q_ycrcb[:, :, Channel].astype(np.float32)
                    size = block_radius + 1
                    A = np.mean(Y_block[0:size, 0:size])
                    B = np.mean(Y_block[0:size, -size:])
                    C = np.mean(Y_block[-size:, 0:size])
                    D = np.mean(Y_block[-size:, -size:])
                    digit = E4[wrong_index]
                    print(f"提取x,y为{x, y},vals为A={A:.2f}, B={B:.2f}, C={C:.2f}, D={D:.2f}，digit为{digit}")

                wrong_index += 1

            print(f"{mode_name}_不同字符的百分比: {diff_percentage:.2f}%")


            def binary_to_image(bstr, shape):
                array = np.array(list(map(int, bstr)), dtype=np.uint8).reshape((shape, shape))
                return array * 255  # 黑白图


            extracted_wm = binary_to_image(E2, watermark_shape)


            def normalized_correlation(w_original, w_extracted):
                w_original = w_original.flatten().astype(np.float32)
                w_extracted = w_extracted.flatten().astype(np.float32)

                numerator = np.sum(w_original * w_extracted)
                denominator = np.sum(w_original ** 2) + 1e-6  # 防除0
                nc = numerator / denominator
                return nc


            def bit_error_rate(w_original, w_extracted):
                w_original = w_original.flatten().astype(np.uint8)
                w_extracted = w_extracted.flatten().astype(np.uint8)

                total_bits = len(w_original)
                error_bits = np.sum(w_original != w_extracted)
                return error_bits / total_bits


            pixels = np.array([int(char) * 255 for char in E2], dtype=np.uint8)
            # 重新调整为 watermark_shape x watermark_shape 的二维数组
            watermark_extract = pixels.reshape((watermark_shape, watermark_shape))
            # 保存为 PNG 灰度图像
            # cv2.imwrite("watermark_extract.png", watermark_extract)
            cv2.imwrite(f"{output_folder}/{mode_name}_watermark_extract.png", watermark_extract)
            extracted_water = cv2.imread(f"{output_folder}/{mode_name}_watermark_extract.png")

            nc = normalized_correlation(water_i, extracted_water)
            ber = bit_error_rate(water_i, extracted_water)
            print(f"{mode_name}_图像PSNR 值: {psnr_image:.2f} dB")
            print(f"{mode_name}_图像SSIM 值: {ssim_image:.4f} ")
            print("  ")
            print(f"NC: {nc:.4f}")
            print(f"BER: {ber:.4f}")
            row_data.extend([nc, ber])  # 或者 row_data += [nc, ber]
            print(
                f"not_one为{one_not_enough},not_two为{two_not_enough},four为{four_not_enough},,six为{six_not_enough},eight={eight_not_enough},get_nothing_extract={get_nothing_extract}")
            print(f"type4={type4},type5={type5},type6={type6}")
            print(f"最小值为{global_min},最大值为{global_max}")

            # 恢复标准输出，并关闭日志文件
        sys.stdout = sys.__stdout__
        log_file.close()

        ws.append(row_data)

        # 空一行作为分隔
        start_row = ws.max_row + 2
        ws.append([])
        wb.save(excel_path)

print("All processing finished.")

wb.save(excel_path)







